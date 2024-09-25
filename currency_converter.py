from tkinter import Tk,ttk
from tkinter import *
from tkinter import filedialog
from tkinter import messagebox
import time
import PIL
from PIL import ImageTk,Image
from PIL import *
import openpyxl
from openpyxl.styles import Border,Side
import pandas as pd
import xlwings as xw
from xlwings import constants
import xlsxwriter
from pathlib import Path
import win32com.client
from pywintypes import com_error

import selenium
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys

import logging
import getpass
from datetime import datetime
#Configure the logging
logging.basicConfig(filename="app.log",level=logging.INFO)

#Get the current username
username = getpass.getuser()

#Function to format log messages with username and timestamp
def format_log_message(message):
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    return f"[User:{username}][{timestamp}]{message}"

#colors
cor0 = '#1A3180'
cor1 = '#FDA1E6'
cor2 = '#91E1F1'


root = Tk()
root.geometry("800x500")
root.title('Currency Converter')
root.configure(bg=cor0)
root.resizable(height=True,width=True)

top = Frame(root,width=700,height=200,bg=cor0)
top.grid(row=0,column=0)

#Buttons
# Clear_Data_btn = Button(top,text='Clear Data',width=12,height=3,command=Clear_Data, font=('arial 10 bold'),bg=cor1,fg=cor0)
# Clear_Data_btn.place(x=150,y=22)

# Export_to_PDF_btn = Button(top,text='Export to PDF',width=12,height=3,wraplength=65,command=Export_pdf,font=('arial 10 bold'),bg=cor1,fg=cor0)
# Export_to_PDF_btn.place(x=450,y=22)

Export_to_Excel_btn = Button(top,text='Export to Excel',width=12,height=3,wraplength=65,font=('arial 10 bold'),bg=cor2,fg=cor0)
Export_to_Excel_btn.place(x=570,y=22)

#Lables
Source_currency_lable = Label(top,text='Source currency',width=21,height=1,font=('arial 10'),border=1,bg='blue',fg='white')
Source_currency_lable.place(x=1,y=119)

Currency_Input_Value_lable = Label(top,text='Currency Input Value ',width=21,height=1,font=('arial 10'),border=1,bg='blue',fg='white')
Currency_Input_Value_lable.place(x=175,y=119)

Destination_Currency_lable = Label(top,text='Destination Currency',width=21,height=1,font=('arial 10'),border=1,bg='blue',fg='white')
Destination_Currency_lable.place(x=350,y=119)

Converted_Currency_lable = Label(top,text='Converted Currency',width=21,height=1,font=('arial 10'),border=1,bg='green',fg='white')
Converted_Currency_lable.place(x=525,y=119)

#ComboBox
combo1 = ttk.Combobox(width=21,height=1,justify=CENTER,font=('arial 10'))
combo1.place(x=1,y=140)

strlist = ['']
workbook = openpyxl.load_workbook("E:\python\Currency_list.xlsx")
wkslist = workbook['Sheet1']
lastrow = wkslist.max_row

#don't remember the method name
lastrow = int(lastrow)
for row1 in range(1,lastrow):
    row1 = int(row1)
    strlist.append(wkslist.cell(row1,1).value)
combo1['values'] = strlist


combo2 = ttk.Combobox(width=21,height=1,justify=CENTER,font=('arial 10'))
combo2.place(x=350,y=140)

strlist = ['']
workbook = openpyxl.load_workbook("E:\python\Currency_list.xlsx")
wkslist = workbook['Sheet1']
lastrow = wkslist.max_row

#don't remember the method name
lastrow = int(lastrow)
for row1 in range(1,lastrow):
    row1 = int(row1)
    strlist.append(wkslist.cell(row1,1).value)
combo2['values'] = strlist


#Currency input value
value = Text(root,width=24,height=1,font=('arial 10'))
value.place(x=175,y=140)

#Converted currency
converted_currency = Entry(width=24,font=('arial 10'))
converted_currency.place(x=525,y=140)


#Bulk Input
bulk_input = Label(text='Bulk Input',width=22,height=1,justify='left',bg='green',fg='white')
bulk_input.place(x=1,y=180)

filepath_lable = Entry(root,width=50,justify=CENTER)
filepath_lable.place(x=175,y=180)


#Browse button
browse_btn = Button(text="Browse",height=1,width=18,bg='white',command=lambda:browse_button(),font=("times new roman",9))
browse_btn.place(x=550,y=180)

#check values
convert_currency_btn = Button(top,text='Convert Currency',width=12,height=3,wraplength=58,command=lambda:convert_currency(),font=('arial 10 bold'),bg=cor1,fg=cor0)
convert_currency_btn.place(x=25,y=22)



def convert_currency():
    # service = Service()
    # options = webdriver.ChromeOptions()
    driver = webdriver.Chrome("E:\Drivers\chromedriver.exe")
    driver.get('https://www.google.com/search?q=google+currency+converter')
    driver.maximize_window()
    time.sleep(5)
    
    
    #get the input values
    s1 = combo1.get()
    d1 = combo2.get()
    val1 = value.get("1.0",END)
    converted_currency.delete(0,END)
    print(s1)
    print(val1)
    print(d1)
    


    #source value
    source_value = driver.find_element(By.XPATH,'//input[@class="lWzCpb ZEB7Fb"]')
    source_value.click()
    source_value.send_keys(1*Keys.BACK_SPACE)
    source_value.send_keys(val1)
    
    #Destination country
    destination_search = driver.find_element(By.XPATH,'//select[@class="zuzy3c NKvwhd"]')
    destination_search.send_keys(d1)
    time.sleep(10)
    
    #source country
    source_search = driver.find_element(By.XPATH,'//select[@class="zuzy3c l84FKc"]')
    source_search.send_keys(s1)
    time.sleep(5)
    
    #Get the values from web
    dest_value = driver.find_element(By.XPATH,'//input[@class="lWzCpb a61j6"]')
    a = dest_value.get_attribute("value")
    print("value",a)
    print("***************")
    converted_currency.insert(1,a)
    
    
#Clear data
def Clear_Data():
    combo1.delete(0,'end')
    combo2.delete(0,'end')
    value.delete(1.0,'end')
    converted_currency.delete(0,'end')
    filepath_lable.delete(0,'end')
 
Clear_Data_btn = Button(top,text='Clear Data',width=12,height=3,command=Clear_Data, font=('arial 10 bold'),bg=cor1,fg=cor0)
Clear_Data_btn.place(x=150,y=22)
       
#define browse button
def browse_button():
    file_path = filedialog.askopenfile(mode="r",filetypes=[('Excel Files','*.xlsx')])
    if file_path:
        print("File selected:",file_path)
        filepath_lable.insert(0,f'{file_path.name}')
    else:
        messagebox.showinfo("Information","You have not selected any file")


# def apply_borders(wb):
#     try:
#         wb = openpyxl.load_workbook("E:\python\converted currency.xlsx")
#         ws = wb.active
        
#         row_cnt = ws.max_row
#         for i in range(1,row_cnt+1):
#             for col_num in range(1,4):
#                 cell = ws.cell(row=i,column=col_num)
#                 cell.border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
        
    # except Exception as e:
    #     error_message = f"an error occurred while applying borders:{str(e)}"
    #     logging.error(format_log_message(error_message))
    #     messagebox.showinfo("Error",f"an error occurred:{str(e)}")        

def convert():
    input_file = filepath_lable.get()
    print(input_file)
    try:
        if input_file == "":
            messagebox.showinfo("information","Please select the input file")
            #exit()
        wb = xw.Book(input_file)
        ws = wb.sheets[0]
        
        service = Service()
        options = webdriver.ChromeOptions()
        driver = webdriver.Chrome(service=service, options=options)
        #driver = webdriver.Chrome("E:\Drivers\chromedriver.exe")

        driver.get('https://www.google.com/search?q=google+currency+converter')
        driver.maximize_window()
        time.sleep(2)
        Row_cnt = ws.range('A'+str(ws.cells.last_cell.row)).end('up').row
        for i in range(2,Row_cnt+1):
            from_currency = ws.cells(i,1).value
            amount = ws.cells(i,2).value
            to_currency = ws.cells(i,3).value
           
            
            #find input fields
            from_input = driver.find_element(By.XPATH,'//select[@class="zuzy3c l84FKc"]')
            from_input.send_keys(from_currency)
            time.sleep(1)
            print(from_currency)
            
            amount_input = driver.find_element(By.XPATH,'//input[@class="lWzCpb ZEB7Fb"]')
            amount_input.clear()
            amount_input.send_keys(amount)
            amount_input.click()
            time.sleep(1)
            print(amount)
            
            
            to_input = driver.find_element(By.XPATH,'//select[@class="zuzy3c NKvwhd"]')
            to_input.send_keys(to_currency)
            time.sleep(1)
            print(to_currency)
            
            convert_value = driver.find_element(By.XPATH,"//input[@class='lWzCpb a61j6']")
            a = convert_value.get_attribute("value")
            ws.cells(i,4).value = a
            time.sleep(1)
            print(a)
            print("************")
            
            #Log Success
            success_message = "Conersion completed successfully"
            logging.info(format_log_message(success_message))
                   
    except Exception as e:
        #Handle the error and log it
        error_message = f"An error occured during conversion:{str(e)}"
        logging.error(format_log_message(error_message))
        messagebox.showinfo("Error",f"An error occured:{str(e)}")
    
    
    driver.close()

#convert button 
convert_btn = Button(text="convert",width=15,height=1,bg='yellow',fg='black',command=convert,font=("times new roman",9))
convert_btn.place(x=300,y=210)   

     
def Export_pdf():
    input_file = filepath_lable.get()
    try:
        xlapp = win32com.client.Dispatch("Excel.Application")
        xlapp.Visible = True
        wb = xlapp.Workbooks.Open(input_file)
        ws_indes_list = [1]
        wb.Worksheets(ws_indes_list).Select()
        wb.ActiveSheet.ExportAsFixedFormat(0,"E:\test.pdf")
        print("Success")
        
        #Log Success
        logging.info(format_log_message("PDF export completed successfully."))
    
    except Exception as e:
        #Handle the error and log it
        logging.error(format_log_message(f"An error occurred during PDF export:{str(e)}"))
        messagebox.showinfo("Error",f"An error occured:{str(e)}")
    
Export_to_PDF_btn = Button(top,text='Export to PDF',width=12,height=3,wraplength=65,command=Export_pdf,font=('arial 10 bold'),bg=cor1,fg=cor0)
Export_to_PDF_btn.place(x=450,y=22)

root.mainloop()
